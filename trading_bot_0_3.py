# trading_bot_0_3.py
# -*- coding: utf-8 -*-
"""
v7_plus (موثوق) — نسخة أحادية الملف لِـ Render Web Service

- FastAPI لربط المنفذ + Health
- Keep-Alive لمنع spin-down على الخطة المجانية
- لوب المسح على دفعات + Backoff تلقائي لِـ OKX (50011)
- إستراتيجية v7_plus:
  EMA20/50 + حجم ≥ VOL_SPIKE_MIN × المتوسط + ADX ≥ ADX_MIN + منع مطاردة (امتداد ≤ EXTENSION_MAX_ATR × ATR)
  SL = max(1.25×ATR, خلف EMA50 بـ 0.35×ATR)
  TP = 1.0×, 2.0×, 3.2× ATR، وبعد TP1 نحرك SL لـ BE، وبعد TP2 نتتبّع EMA20
- تقارير Heartbeat/Diag على التليجرام
"""

import os, time, math, random, traceback, threading
from datetime import datetime, timezone
from typing import Dict, Any, List, Tuple, Optional

import requests
import numpy as np
import pandas as pd
from fastapi import FastAPI
import uvicorn
import ccxt
from openpyxl import Workbook, load_workbook

# ================== إعدادات (مضمّنة كما طلبت) ==================
PORT = int(os.getenv("PORT", "10000"))

# Telegram (مضمّن)
TG_TOKEN   = "8130568386:AAGmpxKQw1XhqNjtj2OBzJ_-e3_vn0FE5Bs"
TG_CHAT_ID = 8429537293
TG_API     = f"https://api.telegram.org/bot{TG_TOKEN}"

# Keep-Alive
PUBLIC_URL = "https://trading-bot-6arv.onrender.com"  # رابط خدمتك على Render
KEEPALIVE_SEC = 240  # كل 4 دقائق

# Exchange = OKX (Binance/Bybit محجوبين على Render)
EXCHANGE_ID = "okx"
TIMEFRAME   = "5m"
OHLCV_LIMIT = 500

# المسح على دفعات لتخفيف الضغط
SCAN_TOP     = 120     # أعلى 120 رمزا بالسيولة
SCAN_BATCH   = 35      # كم رمز في كل دورة
CYCLE_SLEEP  = 30      # ثواني بين الدورات

# حدود OKX للتعامل مع 50011
OKX_SLEEP_MS       = 1200     # نوم بعد كل نداء عام ~ 1.2s
OKX_MAX_RETRIES    = 4
OKX_BACKOFF_FACTOR = 1.8

# شروط الاستراتيجية (قابلة للضبط)
VOL_SPIKE_MIN      = 1.1
ADX_MIN            = 18.0
EXTENSION_MAX_ATR  = 1.2
MIN_RR             = 1.18

ATR_MULT_TP        = (1.0, 2.0, 3.2)
SL_ATR_BASE        = 1.25
SL_BEHIND_EMA50_ATR= 0.35
COOLDOWN_HOURS     = 3
MIN_DOLLAR_VOLUME  = 800_000   # حد أدنى للسيولة
MAX_SPREAD_PCT     = 0.20/100  # 0.20%

# تقارير
DIAG_EACH_SEC      = 600       # كل 10 دقائق
HEARTBEAT_SEC      = 1800      # كل 30 دقيقة

# ================== Telegram ==================
def tg(text: str):
    try:
        requests.post(f"{TG_API}/sendMessage",
                      json={"chat_id": TG_CHAT_ID, "text": text, "parse_mode": "HTML"},
                      timeout=15)
    except Exception:
        pass

# ================== ملفات Excel ==================
BASE_DIR = os.path.dirname(__file__)
SIG_XLSX = os.path.join(BASE_DIR, "signals_v7_plus.xlsx")
REJ_XLSX = os.path.join(BASE_DIR, "reject_v7_plus.xlsx")

def init_excels():
    try:
        if not os.path.exists(SIG_XLSX):
            wb = Workbook(); ws = wb.active; ws.title = "Signals"
            ws.append(["#", "Pair", "Side", "Entry", "TP1", "TP2", "TP3", "SL",
                       "Result", "EntryTime", "ExitTime", "DurationMin", "ProfitPct",
                       "EntryReason", "ExitReason"])
            wb.save(SIG_XLSX)
        if not os.path.exists(REJ_XLSX):
            wb = Workbook(); ws = wb.active; ws.title = "Rejects"
            ws.append(["#", "Pair", "Reason", "ADX", "VolSpike", "ExtATR", "Spread%", "DVol", "Time"])
            wb.save(REJ_XLSX)
    except Exception:
        pass

def xl_append_signal(sig: dict, res: str, exit_t: str, dur_min: float, pl_pct: float, exit_rsn: str):
    try:
        wb = load_workbook(SIG_XLSX); ws = wb.active
        ws.append([ws.max_row,
                   sig["symbol"].replace("/USDT:USDT", "/USDT"),
                   sig["side"], sig["entry"], *sig["tps"], sig["sl"],
                   res, sig["start_time"], exit_t, round(dur_min,1), round(pl_pct,2),
                   " • ".join(sig.get("reasons", [])[:4]), exit_rsn])
        wb.save(SIG_XLSX)
    except Exception:
        pass

def xl_append_reject(sym: str, reason: str, adx: float, vsp: float, ext: float, spr: float, dvol: float):
    try:
        wb = load_workbook(REJ_XLSX); ws = wb.active
        ws.append([ws.max_row, sym.replace("/USDT:USDT","/USDT"), reason,
                   round(adx,1), round(vsp,2), round(ext,2), round(spr*100,3),
                   round(dvol,0), datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")])
        wb.save(REJ_XLSX)
    except Exception:
        pass

# ================== مؤشرات فنية ==================
def ema(s, n): return s.ewm(span=n, adjust=False).mean()

def rsi(s, n=14):
    d = s.diff(); up = np.maximum(d, 0); dn = np.maximum(-d, 0)
    rs = (pd.Series(up).ewm(span=n, adjust=False).mean() /
          (pd.Series(dn).ewm(span=n, adjust=False).mean() + 1e-12))
    return 100 - 100 / (1 + rs)

def tr(h, l, c):
    pc = c.shift(1)
    return pd.concat([h-l, (h-pc).abs(), (l-pc).abs()], axis=1).max(axis=1)

def atr(h, l, c, n=14): return tr(h,l,c).ewm(span=n, adjust=False).mean()

def adx(h, l, c, n=14):
    up = h.diff(); dn = -l.diff()
    up[up < 0] = 0; dn[dn] < 0
    trv = tr(h,l,c); atrv = trv.ewm(span=n, adjust=False).mean()
    pdi = 100 * (up.ewm(span=n, adjust=False).mean() / (atrv + 1e-12))
    mdi = 100 * (dn.ewm(span=n, adjust=False).mean() / (atrv + 1e-12))
    dx  = 100 * (abs(pdi - mdi) / (pdi + mdi + 1e-12))
    return dx.ewm(span=n, adjust=False).mean(), pdi, mdi

# ================== أكسشينج ==================
def new_exchange():
    return getattr(ccxt, EXCHANGE_ID)({
        "enableRateLimit": True,
        "timeout": 20000,
        "options": {
            "defaultType": "swap",  # لكي نحصل على /USDT:USDT
        },
    })

ex = None

def okx_sleep(ms: int):
    time.sleep(max(ms, 200)/1000.0)

def is_rate_limited(e: Exception) -> bool:
    s = str(e)
    return "Too Many Requests" in s or "50011" in s or "RateLimit" in s

# ================== ماركت/داتا ==================
def safe_fetch(fn, *args, **kwargs):
    """لفّافة نداءات عامة مع Backoff لِـ OKX."""
    delay = OKX_SLEEP_MS
    for i in range(OKX_MAX_RETRIES):
        try:
            res = fn(*args, **kwargs)
            okx_sleep(delay)
            return res
        except Exception as e:
            if is_rate_limited(e):
                # backoff
                delay = int(delay * OKX_BACKOFF_FACTOR)
                time.sleep(min(6, 0.5 + i))  # تهدئة بسيطة
                continue
            else:
                raise
    # آخر محاولة بدون raise
    try:
        res = fn(*args, **kwargs)
        okx_sleep(delay)
        return res
    except Exception as e:
        print(f"ERROR: {fn.__name__} failed after retries: {e}")
        return None

def load_markets_and_top_symbols() -> List[str]:
    """تحميل الأسواق مرة ثم ترتيب أعلى USDT سواب حسب السيولة (quoteVolume)."""
    try:
        mkts = safe_fetch(ex.load_markets)
        if not mkts:
            return []
        syms = [s for s, d in mkts.items()
                if (d.get("type") == "swap" or d.get("swap"))
                and d.get("quote") in ("USDT", "USDT:USDT")
                and d.get("active")]
        ticks = safe_fetch(ex.fetch_tickers) or {}
        syms.sort(key=lambda s: float((ticks.get(s) or {}).get("quoteVolume") or 0), reverse=True)
        return syms[:SCAN_TOP]
    except Exception as e:
        print("load_markets error:", e)
        return []

def ohlcv(sym: str, lim=OHLCV_LIMIT) -> Optional[pd.DataFrame]:
    try:
        data = safe_fetch(ex.fetch_ohlcv, sym, TIMEFRAME, limit=lim)
        if not data or len(data) < 200:
            return None
        df = pd.DataFrame(data, columns=["ts","o","h","l","c","v"])
        df["ts"] = pd.to_datetime(df["ts"], unit="ms", utc=True)
        for col in ["o","h","l","c","v"]:
            df[col] = df[col].astype(float)
        return df
    except Exception as e:
        print("ohlcv error", sym, e)
        return None

def last_price(sym: str) -> Optional[float]:
    try:
        t = safe_fetch(ex.fetch_ticker, sym)
        return float(t["last"]) if t else None
    except Exception:
        return None

def spread_pct(sym: str) -> float:
    try:
        ob = safe_fetch(ex.fetch_order_book, sym, 5) or {}
        bids = ob.get("bids") or []
        asks = ob.get("asks") or []
        if not bids or not asks: return 1.0
        bid, ask = bids[0][0], asks[0][0]
        return (ask - bid) / ask if ask else 1.0
    except Exception:
        return 1.0

def dollar_vol(df: pd.DataFrame) -> float:
    try:
        return float((df["c"].iloc[-30:] * df["v"].iloc[-30:]).sum())
    except Exception:
        return 0.0

# ================== الإستراتيجية ==================
def crossed_recent(f: pd.Series, s: pd.Series, side: str) -> Tuple[bool, Optional[int]]:
    rng = range(1, 7)
    for i in rng:
        prev_ok = f.iloc[-i-1] <= s.iloc[-i-1] if side=="LONG" else f.iloc[-i-1] >= s.iloc[-i-1]
        now_ok  = f.iloc[-i]   >  s.iloc[-i]   if side=="LONG" else f.iloc[-i]   <  s.iloc[-i]
        if prev_ok and now_ok:
            return True, i
    return False, None

def strat(sym: str, df: pd.DataFrame):
    c,h,l,v = df["c"], df["h"], df["l"], df["v"]
    last = float(c.iloc[-1])
    e20, e50 = ema(c,20), ema(c,50)
    atr14 = float(atr(h,l,c).iloc[-1])
    adx14, _, _ = adx(h,l,c)
    adx_cur = float(adx14.iloc[-1])
    vsp = float(v.iloc[-1] / (v.rolling(20).median().iloc[-1] + 1e-12))

    if e20.iloc[-1] > e50.iloc[-1]:
        side = "LONG"
    elif e20.iloc[-1] < e50.iloc[-1]:
        side = "SHORT"
    else:
        xl_append_reject(sym, "no_trend", adx_cur, vsp, 0.0, 0.0, 0.0)
        return None, {"reasons":["no_trend"], "atr":atr14}

    cross_ok, cross_age = crossed_recent(e20, e50, side)
    if not cross_ok:
        xl_append_reject(sym, "no_recent_cross", adx_cur, vsp, 0.0, 0.0, 0.0)
        return None, {"reasons":["no_recent_cross"], "atr":atr14}

    if vsp < VOL_SPIKE_MIN:
        xl_append_reject(sym, f"low_vol_{vsp:.1f}x", adx_cur, vsp, 0.0, 0.0, dollar_vol(df))
        return None, {"reasons":["low_volume"], "atr":atr14}

    if adx_cur < ADX_MIN:
        xl_append_reject(sym, f"adx_{int(adx_cur)}", adx_cur, vsp, 0.0, 0.0, dollar_vol(df))
        return None, {"reasons":["weak_adx"], "atr":atr14}

    ext_atr = abs(last - e20.iloc[-1]) / (atr14 + 1e-12)
    if ext_atr > EXTENSION_MAX_ATR:
        xl_append_reject(sym, "extended_from_ema20", adx_cur, vsp, ext_atr, 0.0, dollar_vol(df))
        return None, {"reasons":["extended"], "atr":atr14}

    # شمعة تأكيد باتجاه الترند
    if side=="LONG" and not (c.iloc[-1] > c.iloc[-2]):
        xl_append_reject(sym, "no_confirm_long", adx_cur, vsp, ext_atr, 0.0, dollar_vol(df))
        return None, {"reasons":["no_conf_long"], "atr":atr14}
    if side=="SHORT" and not (c.iloc[-1] < c.iloc[-2]):
        xl_append_reject(sym, "no_confirm_short", adx_cur, vsp, ext_atr, 0.0, dollar_vol(df))
        return None, {"reasons":["no_conf_short"], "atr":atr14}

    r = float(rsi(c).iloc[-1])
    if (side=="LONG" and r>=70) or (side=="SHORT" and r<=30):
        xl_append_reject(sym, f"rsi_extreme_{int(r)}", adx_cur, vsp, ext_atr, 0.0, dollar_vol(df))
        return None, {"reasons":["rsi_extreme"], "atr":atr14}

    reasons = [f"ema20_50_cross_{cross_age}c", f"vol_spike_{vsp:.1f}x", f"adx_{int(adx_cur)}", "confirm_candle", "rsi_ok"]
    return side, {"atr":atr14, "adx":adx_cur, "vsp":vsp, "reasons":reasons}

def build_sl(entry: float, side: str, atr_val: float, ema50_last: float) -> float:
    if side=="LONG":
        sl = min(entry - SL_ATR_BASE*atr_val, ema50_last - SL_BEHIND_EMA50_ATR*atr_val)
    else:
        sl = max(entry + SL_ATR_BASE*atr_val, ema50_last + SL_BEHIND_EMA50_ATR*atr_val)
    return round(sl, 6)

def targets(entry: float, atr_val: float, side: str) -> List[float]:
    m1, m2, m3 = ATR_MULT_TP
    if side=="LONG":
        t1 = entry + m1*atr_val; t2 = entry + m2*atr_val; t3 = entry + m3*atr_val
    else:
        t1 = entry - m1*atr_val; t2 = entry - m2*atr_val; t3 = entry - m3*atr_val
    return [round(t1,6), round(t2,6), round(t3,6)]

def entry_window_ok() -> bool:
    now = datetime.now(timezone.utc)
    sec = (now.minute % 5)*60 + now.second
    return 30 <= sec <= 150

# ================== إدارة الحالات ==================
cooldown: Dict[str, float] = {}
open_trades: Dict[str, dict] = {}
reject_counter: Dict[str, int] = {}
last_diag = 0.0
last_hb   = 0.0

def incr_rej(key: str):
    reject_counter[key] = reject_counter.get(key, 0) + 1

def scan_symbols_batch(all_syms: List[str], start_idx: int) -> Tuple[List[dict], int]:
    picks = []
    n = len(all_syms)
    if n == 0: return picks, start_idx
    end_idx = min(n, start_idx + SCAN_BATCH)
    batch = all_syms[start_idx:end_idx]
    next_idx = 0 if end_idx >= n else end_idx

    for sym in batch:
        try:
            if cooldown.get(sym, 0) > time.time():
                incr_rej("cooldown")
                continue

            df = ohlcv(sym, OHLCV_LIMIT)
            if df is None:
                incr_rej("ohlcv_fail")
                continue

            dvol = dollar_vol(df)
            if dvol < MIN_DOLLAR_VOLUME:
                incr_rej("low_dvol")
                continue

            spr = spread_pct(sym)
            if spr > MAX_SPREAD_PCT:
                incr_rej("high_spread")
                continue

            lp = last_price(sym)
            if not lp:
                incr_rej("no_price")
                continue

            side, meta = strat(sym, df)
            if side is None:
                reason = (meta.get("reasons") or ["unknown"])[0]
                incr_rej(reason)
                continue

            e50 = float(ema(df["c"],50).iloc[-1])
            sl  = build_sl(lp, side, meta["atr"], e50)
            tps = targets(lp, meta["atr"], side)

            rr = abs((tps[0]-lp)/(sl-lp)) if (sl-lp)!=0 else 0
            if rr < MIN_RR:
                incr_rej("poor_rr")
                continue

            picks.append({
                "symbol": sym, "side": side, "entry": float(lp),
                "tps": tps, "sl": float(sl), "atr": meta["atr"],
                "reasons": meta["reasons"], "start_time": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"),
                "state": "OPEN", "tp_stage": 0
            })
        except Exception as e:
            incr_rej("scan_err")
            print("scan error", sym, e)
    return picks, next_idx

def send_signal(sig: dict):
    emoji = "🚀" if sig["side"]=="LONG" else "🔻"
    txt = (f"<b>⚡ توصية v7+</b>\n"
           f"{emoji} <b>{sig['side']}</b> <code>{sig['symbol'].replace('/USDT:USDT','/USDT')}</code>\n\n"
           f"💰 Entry: <code>{sig['entry']}</code>\n"
           f"🎯 TP1: <code>{sig['tps'][0]}</code> | TP2: <code>{sig['tps'][1]}</code> | TP3: <code>{sig['tps'][2]}</code>\n"
           f"🛑 SL: <code>{sig['sl']}</code>\n\n"
           f"🔎 Signals: <i>{' • '.join(sig['reasons'][:4])}</i>\n"
           f"🛡️ Risk: SL→BE بعد TP1 • تتبع EMA20 بعد TP2")
    tg(txt)
    open_trades[sig["symbol"]] = sig
    cooldown[sig["symbol"]] = time.time() + COOLDOWN_HOURS*3600

def res_exit(sym: str, st: dict, result: str, price: float, exit_reason: str):
    exit_t = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    dur = (datetime.strptime(exit_t, "%Y-%m-%d %H:%M")
           - datetime.strptime(st["start_time"], "%Y-%m-%d %H:%M")).total_seconds()/60
    if st["side"]=="LONG":
        pl_pct = (price - st["entry"]) / st["entry"] * 100
    else:
        pl_pct = (st["entry"] - price) / st["entry"] * 100
    xl_append_signal(st, result, exit_t, dur, pl_pct, exit_reason)
    emoji = {"SL":"🛑", "TP1":"✅", "TP2":"🎉", "TP3":"🏆"}.get(result, "✅")
    tg(f"{emoji} <b>{result}</b> {sym.replace('/USDT:USDT','/USDT')} @ <code>{price}</code>\n"
       f"P/L: <b>{pl_pct:+.2f}%</b> • مدة: {dur:.0f}m\n"
       f"خروج: {exit_reason}")
    st["state"] = "CLOSED"

def track_trades_once():
    for sym, st in list(open_trades.items()):
        if st.get("state") != "OPEN":
            continue
        lp = last_price(sym)
        if not lp:
            continue

        side = st["side"]; tps = st["tps"]; sl = st["sl"]; entry = st["entry"]
        # EMA20 للتتبّع (خفيفة: 120 شمعة فقط)
        df = ohlcv(sym, 120)
        e20 = float(ema(df["c"],20).iloc[-1]) if df is not None else entry

        if side=="LONG":
            if lp <= sl:
                res_exit(sym, st, "SL", lp, "Stop Loss")
                continue
            if st["tp_stage"] < 1 and lp >= tps[0]:
                st["tp_stage"] = 1
                st["sl"] = entry  # BE
                tg(f"✅ TP1 • SL→BE\nالسعر: <code>{lp}</code>\n{sym.replace('/USDT:USDT','/USDT')}")
            elif st["tp_stage"] < 2 and lp >= tps[1]:
                st["tp_stage"] = 2
                st["sl"] = max(st["sl"], e20)
                tg(f"🎯 TP2 • تتبع SL عند EMA20≈<code>{st['sl']:.6f}</code>\n{sym.replace('/USDT:USDT','/USDT')}")
            elif lp >= tps[2]:
                res_exit(sym, st, "TP3", lp, "Target 3 🎯")
        else:
            if lp >= sl:
                res_exit(sym, st, "SL", lp, "Stop Loss")
                continue
            if st["tp_stage"] < 1 and lp <= tps[0]:
                st["tp_stage"] = 1
                st["sl"] = entry  # BE
                tg(f"✅ TP1 • SL→BE\nالسعر: <code>{lp}</code>\n{sym.replace('/USDT:USDT','/USDT')}")
            elif st["tp_stage"] < 2 and lp <= tps[1]:
                st["tp_stage"] = 2
                st["sl"] = min(st["sl"], e20)
                tg(f"🎯 TP2 • تتبع SL عند EMA20≈<code>{st['sl']:.6f}</code>\n{sym.replace('/USDT:USDT','/USDT')}")
            elif lp <= tps[2]:
                res_exit(sym, st, "TP3", lp, "Target 3 🎯")

# ================== تقارير ==================
def send_cycle_diag(batch_done: int, total_syms: int):
    global last_diag
    if time.time() - last_diag < DIAG_EACH_SEC:
        return
    reasons_sorted = sorted(reject_counter.items(), key=lambda x: x[1], reverse=True)
    lines = [f"📋 لماذا لا توجد توصيات (الدورة)",
             f"عتبات: ATRx≤{EXTENSION_MAX_ATR} • ADX≥{ADX_MIN:.1f} • RR≥{MIN_RR} • vol≥{VOL_SPIKE_MIN}×",
             f"دفعة المسح: {batch_done}/{total_syms}"]
    if reasons_sorted:
        lines.append("أكثر أسباب الرفض:")
        for k, v in reasons_sorted[:6]:
            lines.append(f"• {k}: {v}")
    tg("\n".join(lines))
    last_diag = time.time()

def send_heartbeat():
    global last_hb
    if time.time() - last_hb < HEARTBEAT_SEC:
        return
    open_count = sum(1 for s in open_trades.values() if s.get("state")=="OPEN")
    tg(f"💓 Heartbeat\n⏱️ TF: {TIMEFRAME} • 🔝 Top: {SCAN_TOP} • Batch: {SCAN_BATCH}\nصفقات مفتوحة: {open_count}\nDiag كل 10m • HB كل 30m")
    last_hb = time.time()

# ================== اللوب الرئيسي (خيط منفصل) ==================
def bot_loop():
    global ex, reject_counter
    init_excels()
    tg("🤖 <b>Bot Started • v7_plus</b>\n"
       f"⏱️ TF: {TIMEFRAME} • 🔝 Top: {SCAN_TOP} • Batch: {SCAN_BATCH}\n"
       f"🛑 SL: max({SL_ATR_BASE}×ATR, خلف EMA50 {SL_BEHIND_EMA50_ATR}×ATR)\n"
       f"📈 شروط: EMA20/50 + Vol≥{VOL_SPIKE_MIN}× + ADX≥{ADX_MIN:.1f} • RR≥{MIN_RR}")

    ex = new_exchange()
    top_syms = load_markets_and_top_symbols()
    if not top_syms:
        tg("⚠️ خطأ في تحميل قائمة الرموز من OKX — سيعاد المحاولة تلقائيًا.")
    next_idx = 0
    last_refresh = time.time()

    while True:
        try:
            if time.time() - last_refresh > 1800:  # جدّد كل 30 دقيقة
                top_syms = load_markets_and_top_symbols()
                last_refresh = time.time()
                next_idx = 0

            if not top_syms:
                time.sleep(5)
                continue

            # امسح دفعة
            reject_counter = {}
            picks, next_idx = scan_symbols_batch(top_syms, next_idx)

            # نافذة الدخول
            if picks and not entry_window_ok():
                tg("⏳ لسنا في نافذة دخول (0:30–2:30 من الشمعة). تجاهلت إشارات هذه الدورة.")
                picks = []

            # أرسل بحد أقصى إشارتين
            for s in picks[:2]:
                send_signal(s)
                time.sleep(2)

            # تتبّع خفيف لِـ الصفقات المفتوحة خلال الانتظار
            t0 = time.time()
            while time.time() - t0 < CYCLE_SLEEP:
                try:
                    track_trades_once()
                except Exception:
                    pass
                time.sleep(5)

            # تقارير
            send_cycle_diag(min(next_idx or SCAN_TOP, SCAN_TOP), len(top_syms))
            send_heartbeat()

        except Exception as e:
            # لا نطيّح البوت
            print("main loop error:", e, traceback.format_exc())
            time.sleep(5)

# ================== Keep-Alive (خيط منفصل) ==================
def keepalive_loop():
    if not PUBLIC_URL:
        return
    path = "/health"
    while True:
        try:
            requests.get(PUBLIC_URL + path, timeout=10)
        except Exception:
            pass
        time.sleep(KEEPALIVE_SEC)

# ================== FastAPI ==================
app = FastAPI()

@app.get("/")
def root():
    return {
        "status": "ok",
        "time": datetime.now(timezone.utc).isoformat(),
        "tf": TIMEFRAME,
        "scan_top": SCAN_TOP,
        "batch": SCAN_BATCH,
        "open_trades": sum(1 for s in open_trades.values() if s.get("state")=="OPEN"),
    }

@app.get("/health")
def health():
    return {"ok": True}

@app.get("/metrics")
def metrics():
    return {
        "rejects": reject_counter,
        "cooldown_count": len(cooldown),
        "open_trades": {k:v for k,v in open_trades.items() if v.get("state")=="OPEN"},
    }

def start_threads_once():
    # شغّل اللوبات في خيوط منفصلة
    th1 = threading.Thread(target=bot_loop, daemon=True)
    th1.start()
    th2 = threading.Thread(target=keepalive_loop, daemon=True)
    th2.start()

# ================== تشغيل ==================
if __name__ == "__main__":
    start_threads_once()
    uvicorn.run(app, host="0.0.0.0", port=PORT, log_level="info")
